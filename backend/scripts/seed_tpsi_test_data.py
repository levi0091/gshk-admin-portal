"""Seed CR's TPSI TEST accounts, companies and officer associations into DEV.

Deliberately NOT an Alembic migration: migrations run in every environment, and
this data must never reach PROD. Idempotent — safe to re-run.

    cd backend && uv run python scripts/seed_tpsi_test_data.py [--dry-run]

Source: docs/TPSI260727100116.xlsx
  Accounts             3 individual TPSI logins (Director / Secretary / Reserve).
                       There are NO corporate login accounts: "corporate
                       officers" are companies acting as officers, they hold no
                       TPSI credentials, and CR forbids them from PIN-signing.
  Companies            5 test companies (4 local + 1 registered non-HK).
  Individual Officers  officer associations for the individual accounts.
  Corporate Officers   companies acting as officers of other companies.

The officer associations are not optional decoration. CR rejects a signature
from anyone not associated with the company as at the signature date
(ERR_MSG_SIGNATORY_NOT_AUTH), so without them the Block 4 signing flow cannot
be exercised against TEST at all.

Idempotency and the seed marker are the same thing: every row is written with
`vp_source_key = 'tpsi_test:<natural key>'`, which is the only unique
constraint on these tables besides the primary key. Re-running upserts in place;
finding seeded rows later is a `vp_source_key LIKE 'tpsi_test:%'` query.
"""
import os
import sys

from openpyxl import load_workbook

# Ensure `backend/` is importable when run as a script from that directory.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from db.supabase import get_supabase  # noqa: E402
from services.tpsi.config import get_config  # noqa: E402

SOURCE = "../docs/TPSI260727100116.xlsx"
MARKER = "tpsi_test:"

# Substrings that mean "this is production" in a connection string.
_PROD_MARKERS = ("prod", "-prd", "production")

# CR "Capacity" -> officer_role enum.
_ROLE = {
    "director": "director",
    "company secretary": "company_secretary",
    "secretary": "company_secretary",
    "reserve director": "reserve_director",
    "authorised representative": "authorised_rep",
    "authorized representative": "authorised_rep",
}


def assert_safe_environment() -> None:
    """Refuse to run anywhere that might be production.

    Two independent checks, because either one alone can be wrong: TPSI_ENV
    describes which CR environment we talk to, DATABASE_URL describes which
    database we write to, and a misconfigured deploy can mismatch them.
    """
    if get_config().env != "test":
        sys.exit("REFUSING: TPSI_ENV is not 'test'. This script never runs against PROD.")

    dsn = (os.environ.get("DATABASE_URL") or "").lower()
    if any(marker in dsn for marker in _PROD_MARKERS):
        sys.exit(
            "REFUSING: DATABASE_URL looks like production. "
            "Point it at DEV before seeding."
        )


def _rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {
            headers[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(row)
            if i < len(headers)
        }
        if any(record.values()):
            yield record


def _date(value: str) -> str | None:
    """CR writes dd/mm/yyyy; Postgres wants ISO."""
    if not value or "/" not in value:
        return None
    day, month, year = value.split("/")[:3]
    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"


def _upsert_by_source_key(table: str, payload: dict) -> dict:
    """Insert-or-update keyed on vp_source_key, without ON CONFLICT.

    The vp_source_key unique indexes on entities/persons/entity_officers are
    PARTIAL (`WHERE vp_source_key IS NOT NULL`). PostgREST's `on_conflict=`
    cannot target a partial index — it has no way to emit the matching
    `index_where` clause — so an upsert against them fails with 42P10,
    "no unique or exclusion constraint matching the ON CONFLICT
    specification". Select-then-write is the portable equivalent here, and this
    seed is small enough that the extra round trip does not matter.
    """
    supabase = get_supabase()
    key = payload["vp_source_key"]
    existing = (
        supabase.table(table).select("id").eq("vp_source_key", key).execute().data
    )
    if existing:
        row_id = existing[0]["id"]
        supabase.table(table).update(payload).eq("id", row_id).execute()
        return {"id": row_id}
    return supabase.table(table).insert(payload).execute().data[0]


def seed(dry_run: bool = False) -> dict:
    assert_safe_environment()

    book = load_workbook(SOURCE, data_only=True)
    accounts = list(_rows(book["Accounts"]))
    companies = list(_rows(book["Companies"]))
    ind_officers = list(_rows(book["Individual Officers"]))
    corp_officers = list(_rows(book["Corporate Officers"]))

    summary = {
        "accounts": len(accounts),
        "companies": len(companies),
        "individual_officers": len(ind_officers),
        "corporate_officers": len(corp_officers),
    }
    print(f"source: {SOURCE}")
    for key, count in summary.items():
        print(f"  {count:>3} {key.replace('_', ' ')}")

    if dry_run:
        print("dry run - nothing written")
        return summary

    # 1. Companies -> entities.
    entity_by_brn: dict[str, str] = {}
    for company in companies:
        brn = company.get("BRN", "")
        if not brn:
            continue
        name_en = company.get("Company Name (Eng)", "")
        name_zh = company.get("Company Name (Chi)", "")
        row = _upsert_by_source_key(
            "entities",
            {
                    "vp_source_key": f"{MARKER}{brn}",
                    "br_number": brn,
                    "company_name": name_en or brn,
                    "company_name_zh": name_zh or None,
                    "name_language": "bilingual" if (name_en and name_zh) else "english",
                    "status": "live",
                    "incorporation_date": _date(company.get("Date of Incorporation", "")),
                    "incorporation_place": company.get("Place of Incorporation") or "Hong Kong",
            },
        )
        entity_by_brn[brn] = row["id"]

    # 2. The 3 individual TPSI login accounts -> persons.
    #    These are the only parties that can PIN-sign. Their TPSI/e-Service
    #    passwords are NOT stored here — presenter credentials belong in
    #    tpsi_presenter_credentials, set per user through the API.
    person_by_userid: dict[str, str] = {}
    person_by_hkid: dict[str, str] = {}
    for account in accounts:
        user_id = account.get("User ID", "")
        if not user_id:
            continue
        surname = account.get("Surname", "")
        given = account.get("Other Name in English", "")
        row = _upsert_by_source_key(
            "persons",
            {
                    "vp_source_key": f"{MARKER}{user_id}",
                    "full_name": f"{surname}, {given}".strip(", ") or user_id,
                    "surname": surname or None,
                    "given_names": given or None,
                    "full_name_zh": account.get("Name in Chinese") or None,
            },
        )
        person_by_userid[user_id] = row["id"]
        hkid = account.get("HKID", "")
        if hkid:
            person_by_hkid[hkid] = row["id"]

    # 3. Officer associations. Without these, CR rejects every signature.
    linked = 0
    for officer in ind_officers:
        brn = officer.get("BRN", "")
        entity_id = entity_by_brn.get(brn)
        person_id = person_by_hkid.get(officer.get("HKID", ""))
        if not (entity_id and person_id):
            continue
        capacity = officer.get("Capacity", "").strip().lower()
        _upsert_by_source_key(
            "entity_officers",
            {
                "vp_source_key": f"{MARKER}ind:{brn}:{officer.get('HKID')}:{capacity}",
                "entity_id": entity_id,
                "person_id": person_id,
                "party_type": "individual",
                "role": _ROLE.get(capacity, "director"),
                "appointed_date": _date(officer.get("Association Date", "")),
                "is_current": True,
            },
        )
        linked += 1

    for officer in corp_officers:
        brn = officer.get("BRN", "")
        entity_id = entity_by_brn.get(brn)
        if not entity_id:
            continue
        capacity = officer.get("Capacity", "").strip().lower()
        officer_brn = officer.get("Officer BRN", "")
        _upsert_by_source_key(
            "entity_officers",
            {
                "vp_source_key": f"{MARKER}corp:{brn}:{officer_brn}:{capacity}",
                "entity_id": entity_id,
                "person_id": None,
                "party_type": "corporate",
                "corporate_name": officer.get("Company Name (Eng)") or officer_brn,
                "role": _ROLE.get(capacity, "director"),
                "appointed_date": _date(officer.get("Association Date", "")),
                "is_current": True,
            },
        )
        linked += 1

    summary["linked"] = linked

    # ---- 4. Everything CR's workbook does NOT carry ------------------------
    #
    # The workbook has no addresses, no share capital and no secretaries, so a
    # seed built from it alone cannot produce a NAR1 that map_entity will even
    # build, let alone one CR accepts. The 2026-08-21 live run had to patch all
    # of this in memory. Synthesised here instead so the next window is spent
    # testing rather than patching.
    supabase = get_supabase()

    # 4a. One registered office per company, one residence per person.
    def _address(key: str, line1: str, city: str) -> str:
        return _upsert_by_source_key("addresses", {
            "vp_source_key": f"{MARKER}addr:{key}",
            "line1": line1, "line2": "Test Tower", "line3": "1 Test Street",
            "city": city, "country": "Hong Kong", "is_hk_address": True,
        })["id"]

    for brn, entity_id in entity_by_brn.items():
        supabase.table("entities").update(
            {"registered_address_id": _address(f"ro:{brn}", "Flat A, 12/F", "CENTRAL")}
        ).eq("id", entity_id).execute()

    for user_id, person_id in person_by_userid.items():
        supabase.table("persons").update({
            "residential_address_id": _address(
                f"res:{user_id}", "Flat B, 8/F", "WAN CHAI"),
            # selectPersonId is this, and CR rejects a signature from a user id
            # it does not recognise ("Please check selectPersonId field.").
            "eservice_user_id": user_id,
        }).eq("id", person_id).execute()

    # 4b. Identity documents. The FULL number is stored, as it is in Viewpoint;
    #     nar1_mapper derives the partial CR wants for the NAR1 ("T001137" +
    #     check digit 6 -> "T001"). Storing the partial here would conflate the
    #     two and break ND2A, which files the full number.
    identity = 0
    for account in accounts:
        person_id = person_by_userid.get(account.get("User ID", ""))
        hkid = account.get("HKID", "")
        if not (person_id and hkid):
            continue
        check = account.get("HKID Check Digit", "")
        _upsert_by_source_key("person_identity_documents", {
            "vp_source_key": f"{MARKER}id:{account['User ID']}",
            "person_id": person_id, "id_type": "hkid",
            "id_number": f"{hkid}({check})" if check else hkid,
            "is_primary": True,
        })
        identity += 1
    summary["identity_documents"] = identity

    # 4c. Corporate officers need their OWN entity, or the mapper has no
    #     address to file for them. The workbook's "Officer BRN" is a company
    #     seeded above, so the FK can simply be pointed at it.
    corporate_linked = 0
    for officer in corp_officers:
        entity_id = entity_by_brn.get(officer.get("BRN", ""))
        officer_entity_id = entity_by_brn.get(officer.get("Officer BRN", ""))
        if not (entity_id and officer_entity_id):
            continue
        capacity = officer.get("Capacity", "").strip().lower()
        key = f"{MARKER}corp:{officer.get('BRN')}:{officer.get('Officer BRN')}:{capacity}"
        supabase.table("entity_officers").update(
            {"corporate_entity_id": officer_entity_id}
        ).eq("vp_source_key", key).execute()
        corporate_linked += 1
    summary["corporate_entity_links"] = corporate_linked

    # 4d. company_secretaries. entity_officers records the officer register;
    #     the NAR1 secretary block is read from this table, and CR's workbook
    #     expresses secretaries only as officer rows.
    secretaries = 0
    for officer in ind_officers:
        if officer.get("Capacity", "").strip().lower() not in ("company secretary", "secretary"):
            continue
        entity_id = entity_by_brn.get(officer.get("BRN", ""))
        person_id = person_by_hkid.get(officer.get("HKID", ""))
        if not (entity_id and person_id):
            continue
        _upsert_by_source_key("company_secretaries", {
            "vp_source_key": f"{MARKER}sec:{officer.get('BRN')}:{officer.get('HKID')}",
            "entity_id": entity_id, "person_id": person_id,
            "is_gshk": False, "is_current": True,
            "appointed_date": _date(officer.get("Association Date", "")),
        })
        secretaries += 1
    summary["secretaries"] = secretaries

    # 4e. Share capital. NAR1 Schedule 1 must account for every issued share,
    #     so the holding below is deliberately the WHOLE issued capital — an
    #     under-allotted class is a MappingError, not a filing.
    shares = 0
    for brn, entity_id in entity_by_brn.items():
        director = next(
            (person_by_hkid.get(o.get("HKID", "")) for o in ind_officers
             if o.get("BRN") == brn
             and o.get("Capacity", "").strip().lower() == "director"),
            None,
        )
        if not director:
            continue
        class_id = _upsert_by_source_key("share_classes", {
            "vp_source_key": f"{MARKER}class:{brn}",
            "entity_id": entity_id, "class_name": "Ordinary",
            "currency": "HKD", "nominal_value": 1, "votes_per_share": 1,
            "total_issued": 100, "total_paid": 100,
        })["id"]
        _upsert_by_source_key("shareholdings", {
            "vp_source_key": f"{MARKER}holding:{brn}",
            "entity_id": entity_id, "share_class_id": class_id,
            "person_id": director, "party_type": "individual",
            "shares_held": 100, "amount_paid": 100, "is_current": True,
        })
        shares += 1
    summary["share_classes"] = shares

    print(
        f"seeded ({linked} officer associations, {identity} identity documents, "
        f"{secretaries} secretaries, {shares} share classes, "
        f"{corporate_linked} corporate officer links). "
        "Re-running is safe - every write is an upsert."
    )
    return summary


if __name__ == "__main__":
    seed(dry_run="--dry-run" in sys.argv)
