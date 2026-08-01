"""One-off: the CR §7.2 worksheet -> committed nar1_schema.json.

Source: docs/Web Form Example/Worksheet in TPSI API Interface v1.0.14.xlsx,
sheet "NAR1". This is CR's XML parameter contract.

It is NOT backend/docs/tpsi/NAR1_Data_Specification.v1.4.xls, which lists
web-UI field names (S2compName, S11A*, Sch1*, formMode, AccBarcode) that appear
in no XML CR ships. Building from that file would produce a document CR rejects.

Layout: the parameter name sits in ONE of columns A..K and the column index IS
the nesting depth (A=0 -> "submission"). L=Type, M=Mandatory, N=Length,
O=Remark, P=Updated Remark. Row 1 is the header.

Run once, commit the JSON: the API then needs no Excel dependency at runtime and
the schema is diffable in review.

    cd backend && uv run python scripts/gen_nar1_schema.py
"""
import json
import pathlib
import sys
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

SRC = pathlib.Path("../docs/Web Form Example/Worksheet in TPSI API Interface v1.0.14.xlsx")
DST = pathlib.Path("services/tpsi/forms/nar1_schema.json")

NAME_COLS = range(0, 11)  # A..K — column index is the nesting depth
TYPE_COL, MAND_COL, LEN_COL, REMARK_COL = 11, 12, 13, 14  # L, M, N, O

# ---------------------------------------------------------------------------
# Worksheet -> XML reconciliation.
#
# CR's worksheet does not perfectly match CR's own shipped example instances in
# docs/Web Form Example/. Where they disagree, THE EXAMPLES WIN — they are what
# CR's server actually produced and accepted; the worksheet is a specification
# document that has drifted.
#
# Every entry below was found by the cross-check at the bottom of this script
# (every element name used in validate_NAR1(*).xml must exist in the schema),
# and each is verified against those files. Do not add an entry without that
# evidence, and do not hand-edit the generated JSON instead.
# ---------------------------------------------------------------------------

# Straight renames: worksheet spelling -> the spelling CR's XML actually uses.
_NAME_FIXUPS = {
    # Worksheet row 3 says "Eform"; every example says "EForm".
    "Eform": "EForm",
    # Worksheet row 58 (inside indSec) says "tcspNo"; the examples say
    # "indvTcspNo". The corporate sibling is "corpTcspNo" in both, so only the
    # individual variant drifted.
    "tcspNo": "indvTcspNo",
}

# Levels the worksheet flattens away. The worksheet hangs clsOfShares and
# shareHolderGrps directly off "shares", but the examples nest them one level
# deeper inside a repeating <share> element:
#     schedule1/shares/share/{clsOfShares, shareHolderGrps/...}
# Without this, the builder would emit shares' children at the wrong depth and
# could not repeat share classes at all.
_INSERT_LEVEL = {
    # parent name -> name of the repeating wrapper to interpose
    "shares": "share",
}


def _cell(row, i):
    return "" if i >= len(row) or row[i] is None else str(row[i]).strip()


def _apply_inserted_levels(node: dict) -> None:
    """Interpose the repeating wrappers the worksheet omits (see _INSERT_LEVEL)."""
    wrapper_name = _INSERT_LEVEL.get(node["name"])
    if wrapper_name and node["children"]:
        wrapper = {
            "name": wrapper_name,
            "depth": node["depth"] + 1,
            "data_type": "",
            "mandatory": False,
            "max_length": None,
            "remark": f"Repeating element (interposed: absent from the worksheet, "
                      f"present in CR's example instances).",
            "children": node["children"],
        }
        for child in wrapper["children"]:
            child["depth"] += 1
        node["children"] = [wrapper]

    for child in node["children"]:
        _apply_inserted_levels(child)


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"missing {SRC}")

    sheet = load_workbook(SRC, data_only=True)["NAR1"]
    rows = list(sheet.iter_rows(values_only=True))

    root = {
        "name": "submission",
        "depth": -1,
        "data_type": "",
        "mandatory": False,
        "max_length": None,
        "remark": "",
        "children": [],
    }
    stack = [root]
    count = 0

    for row in rows[1:]:  # row 1 is the header
        depth = name = None
        for i in NAME_COLS:
            value = _cell(row, i)
            if value:
                depth, name = i, value
                break
        if not name or name == "submission":
            continue

        raw_len = _cell(row, LEN_COL)
        try:
            max_length = int(float(raw_len)) if raw_len else None
        except ValueError:
            max_length = None

        node = {
            "name": _NAME_FIXUPS.get(name, name),
            "depth": depth,
            "data_type": _cell(row, TYPE_COL),
            "mandatory": _cell(row, MAND_COL).upper().startswith("Y"),
            "max_length": max_length,
            "remark": _cell(row, REMARK_COL),
            "children": [],
        }

        while len(stack) > 1 and stack[-1]["depth"] >= depth:
            stack.pop()
        stack[-1]["children"].append(node)
        stack.append(node)
        count += 1

    _apply_inserted_levels(root)

    # ---- Hard gate: every element CR actually sends must exist in the schema.
    #      A builder generated from a schema that is missing an element cannot
    #      emit that element at all, and the omission would only surface as an
    #      opaque CR rejection. Fail the generation instead.
    names: set[str] = set()

    def collect(node):
        for child in node["children"]:
            names.add(child["name"])
            collect(child)

    collect(root)

    examples = sorted(pathlib.Path("../docs/Web Form Example/validateForm").glob("validate_NAR1*.xml"))
    if not examples:
        raise SystemExit("no validate_NAR1*.xml examples found to cross-check against")

    structural = {"Envelope", "Body", "validateForm", "submission"}
    failures = []
    for path in examples:
        doc = ET.fromstring(path.read_text(encoding="utf-8-sig"))
        used = {e.tag.rsplit("}", 1)[-1] for e in doc.iter()}
        missing = sorted((used - names) - structural)
        print(f"cross-check {path.name}: {'OK' if not missing else 'MISSING ' + ', '.join(missing)}")
        if missing:
            failures.append((path.name, missing))

    if failures:
        raise SystemExit(
            "schema does not cover CR's own examples; fix the parse or add a "
            f"documented fixup — {failures}"
        )

    DST.parent.mkdir(parents=True, exist_ok=True)
    DST.write_text(json.dumps(root, indent=2, ensure_ascii=False), encoding="utf8")
    print(f"wrote {DST} - {count} parameters, {len(names)} distinct names")

    def show(node, indent=0):
        if indent > 2:
            return
        for child in node["children"]:
            print("  " * indent + child["name"])
            show(child, indent + 1)

    show(root)


if __name__ == "__main__":
    sys.exit(main())
