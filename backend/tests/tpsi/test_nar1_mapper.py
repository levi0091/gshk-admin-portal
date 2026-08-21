"""services/tpsi/forms/nar1_mapper.py — G-FlowDesk entity graph -> CR dict.

Pure in, pure out: no Supabase, no CR. The round-trip test at the bottom is the
one that matters -- it diffs against CR's OWN shipped example, which is the
source of truth for shape (spec §5 BE-1).
"""
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

from services.tpsi.forms import cr_vocabularies, nar1, nar1_mapper

#: HKT is a fixed UTC+8 offset with no DST. Computed here, not imported from the
#: module under test, so the test pins the offset rather than echoing it.
HKT = timezone(timedelta(hours=8))

#: CR's shipped examples, committed under tests/fixtures — see that directory's
#: README. They used to be read out of the .gitignore'd docs/ folder, so on a
#: clean checkout the round-trip test below raised FileNotFoundError and the two
#: worksheet transcription tests skipped, i.e. never ran in CI at all.
_FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "cr-examples"

SAMPLE = (
    _FIXTURES / "validateForm"
    / "validate_NAR1(Private Company, Schedule 1).xml"
)

#: CR's own vocabularies, for the tests that check the committed copy of them.
WORKSHEET = _FIXTURES / "Worksheet in TPSI API Interface v1.0.14.xlsx"

#: Hard failure, never a skip. These fixtures are committed, so absence means a
#: broken checkout — and a skipped transcription check is what let the country
#: and capacity tables go unverified on every CI run before this.
_MISSING = [p for p in (SAMPLE, WORKSHEET) if not p.exists()]
if _MISSING:
    raise RuntimeError(
        "CR fixtures missing: " + ", ".join(str(p) for p in _MISSING)
    )

#: The base fixture's secretary is GSHK — a body corporate — and the mapper
#: refuses to invent a Body Corporate capacity for it, because which of CR's 15
#: values GSHK signs under is a business question nobody has answered yet. Tests
#: that are not about the signatory state one explicitly and move on.
BODY_CORPORATE_SIGNATORY = {
    "name": "Get Started HK Limited",
    "capacity": "Director of the Company Secretary (Body Corporate)",
    "person_id": None,
    "date": None,
    "is_corporate": True,
}


def mapped(g, *, year=2026, **kw):
    """map_entity() with the signatory stated, for tests about anything else."""
    kw.setdefault("signatory", BODY_CORPORATE_SIGNATORY)
    return nar1_mapper.map_entity(g, year=year, **kw)


def graph(**over):
    base = {
        "entity": {
            "id": "e1",
            "company_name": "TEST COMPANY LIMITED",
            "company_name_zh": "測試有限公司",
            "br_number": "00000001",
            "incorporation_date": "2020-03-01",
            "registered_address_id": "a1",
        },
        "registered_address": {
            "line1": "Flat A", "line2": "Test Tower", "line3": "1 Test Street",
            "city": "CENTRAL", "state_region": None, "postal_code": None,
            "country": "Hong Kong", "is_hk_address": True,
        },
        "officers": [],
        # Every GSHK-managed company has GSHK as its company secretary, and the
        # secretary is who signs the return (Q-030). Without one there is no
        # signatory and map_entity() rightly refuses to build an unsigned
        # statutory declaration -- see test_a_return_with_nobody_to_sign_it_*.
        "secretaries": [{"is_gshk": True,
                         "secretary_name": "Get Started HK Limited",
                         "tcsp_number": "TC000807", "is_current": True}],
        "share_classes": [],
        "shareholdings": [],
        "persons": {},
        "addresses": {},
        "identity_documents": {},
    }
    base.update(over)
    # CR refuses a NAR1 that files any individual with no partial identity
    # number ("Please input the partial HKID number or partial passport
    # number", verified live 2026-08-21), so a graph whose people carry no
    # documents cannot produce a fileable return. Give every person a valid
    # HKID by default; tests that are ABOUT identity pass their own
    # `identity_documents` and this leaves them alone.
    if "identity_documents" not in over:
        base["identity_documents"] = {
            pid: [{"id_type": "hkid", "id_number": "A123456(7)"}]
            for pid in base["persons"]
        }
    return base


def person(pid="p1", **over):
    p = {
        "id": pid, "full_name": "CHAN TAI MAN", "surname": "CHAN",
        "given_names": "TAI MAN", "full_name_zh": "陳大文",
        "email": "test@cr.gov.hk", "residential_address_id": "a2",
    }
    p.update(over)
    return p


ADDR = {
    "line1": "Flat B", "line2": "Res Tower", "line3": "2 Res Street",
    "city": "CENTRAL", "state_region": None, "postal_code": None,
    "country": "Hong Kong", "is_hk_address": True,
}

#: A corporate party's OWN registered office — deliberately different from both
#: the filing entity's registered address and any individual's residence, so a
#: test can tell which one was filed.
CORP_ADDR = {
    "line1": "Suite 1", "line2": "Corp Tower", "line3": "3 Corp Street",
    "city": "WAN CHAI", "state_region": None, "postal_code": None,
    "country": "Hong Kong", "is_hk_address": True,
}


# ---- scalars ---------------------------------------------------------------

def test_maps_the_core_scalars():
    data = mapped(graph())
    assert data["brNo"] == "00000001"
    assert data["yearAnnualReturn"] == 2026
    assert data["language"] == "E"


def test_omits_only_the_two_non_private_company_fields():
    """The ONLY fields where CR's shipped example beats the worksheet. Both
    worksheet remarks end "(Non Private Company)" and this is a private
    company, so the two sources never actually disagreed."""
    data = mapped(graph())
    for absent in ("dateReturnFrom", "dateReturnTo"):
        assert absent not in data


# ---- the statutory signatory block -----------------------------------------

def test_the_signatory_block_is_emitted():
    """CR's own example carries all four (lines 236-239) and the worksheet marks
    them mandatory -- the two sources agree. nar1.validate() never checks
    `mandatory`, so an unsigned return passes every local gate and fails no
    earlier than CR's server, after the fee is taken.

    An individual secretary signing as "Company Secretary" is the path CR's own
    example shows, and it is derived with no help from the caller."""
    g = graph(
        secretaries=[{"person_id": "p1", "is_gshk": False, "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "hkid", "id_number": "A123456(7)",
                                    "is_primary": True}]},
    )
    data = nar1_mapper.map_entity(g, year=2026)      # derived, not passed in
    assert data["selectCapacityDesc"] == "Company Secretary"
    assert data["selectPersonName"] == "CHAN TAI MAN"
    # Today in Hong Kong, not naive date.today(): the DB server runs UTC, which
    # is the wrong calendar day for eight hours out of every twenty-four.
    assert data["signatoryDate"] == datetime.now(HKT).strftime("%d/%m/%Y")


def test_a_body_corporate_signatory_leaves_selectpersonid_empty():
    """Worksheet remark: "Signatory User ID (Empty if sign by Body Corporate)"."""
    assert "selectPersonId" not in mapped(graph())


def test_an_individual_secretary_signs_with_their_identity_number():
    g = graph(
        secretaries=[{"person_id": "p1", "is_gshk": False, "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "hkid", "id_number": "A123456(7)",
                                    "is_primary": True}]},
    )
    data = nar1_mapper.map_entity(g, year=2026)      # derived, not passed in
    assert data["selectPersonName"] == "CHAN TAI MAN"
    assert data["selectPersonId"] == "A1234567"


def test_the_gshk_secretary_signs_in_preference_to_any_other():
    """Q-030 -- GSHK signs on the client's behalf.

    GSHK is a body corporate, so the derived signer now stops for a capacity
    (see test_a_derived_body_corporate_secretary_...). That the mapper stops
    naming GSHK rather than the individual secretary is what proves the
    preference order still holds."""
    g = graph(
        secretaries=[
            {"person_id": "p1", "is_gshk": False, "is_current": True},
            {"is_gshk": True, "secretary_name": "Get Started HK Limited",
             "tcsp_number": "TC000807", "is_current": True},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(g, year=2026)        # derived, not passed in
    problem = next(p for p in exc.value.problems if p.startswith("signatory "))
    assert "Get Started HK Limited" in problem
    assert "CHAN TAI MAN" not in problem


def test_an_explicit_signatory_overrides_the_derived_one():
    data = nar1_mapper.map_entity(
        graph(), year=2026,
        signatory={"name": "WONG SIU MING", "capacity": "Director",
                   "person_id": "B7654321", "date": "2026-06-01"},
    )
    assert data["selectPersonName"] == "WONG SIU MING"
    assert data["selectCapacityDesc"] == "Director"
    assert data["selectPersonId"] == "B7654321"
    assert data["signatoryDate"] == "01/06/2026"


def test_an_explicit_signatory_of_unstated_kind_still_needs_selectpersonid():
    """`is_corporate` is optional in the signatory contract, so an override that
    omits it carries None — and `None is False` is False, so the missing
    mandatory selectPersonId never fired. Task 8's POST /tpsi/filings/prepare is
    the caller of this path; an unstated kind must default to natural person,
    which is the kind CR requires an id for."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(
            graph(), year=2026,
            signatory={"name": "WONG SIU MING", "capacity": "Director",
                       "person_id": None, "date": None},
        )
    assert any("selectPersonId" in p for p in exc.value.problems)


def test_an_explicit_body_corporate_signatory_needs_no_selectpersonid():
    """Worksheet remark: "Signatory User ID (Empty if sign by Body Corporate)".
    Saying so explicitly is the only way to omit the id without a problem."""
    data = nar1_mapper.map_entity(
        graph(), year=2026,
        signatory={"name": "HOLDCO LIMITED",
                   "capacity": "Director of the Company Secretary "
                               "(Body Corporate)",
                   "person_id": None, "date": None, "is_corporate": True},
    )
    assert data["selectPersonName"] == "HOLDCO LIMITED"
    assert "selectPersonId" not in data


def test_the_committed_capacity_vocabularies_are_crs_worksheet():
    """selectCapacityDesc's remark is "Refer to Capacity sheet for description",
    and there are TWO sheets. Prove the committed copies are transcriptions —
    minus each sheet's trailing "for ND4" section, which is another form."""
    import openpyxl

    wb = openpyxl.load_workbook(WORKSHEET, read_only=True, data_only=True)

    def nar1_rows(sheet):
        values = [str(r[0]).strip() for r in wb[sheet].iter_rows(values_only=True)
                  if r and r[0]]
        values = values[1:]                      # header
        if "for ND4" in values:                  # ND4-only tail
            values = values[:values.index("for ND4")]
        return set(values)

    assert cr_vocabularies.CAPACITY_INDIVIDUAL == nar1_rows("Capacity (Individual)")
    assert cr_vocabularies.CAPACITY_BODY_CORPORATE == nar1_rows(
        "Capacity (Body Coporate)")           # CR's own misspelling of the sheet
    assert len(cr_vocabularies.CAPACITY_INDIVIDUAL) == 5
    assert len(cr_vocabularies.CAPACITY_BODY_CORPORATE) == 15


def test_a_body_corporate_may_not_sign_with_an_individual_capacity():
    """"Company Secretary" is in the Individual sheet ONLY. A body corporate
    does not sign — a natural person signs on its behalf, which is what every
    Body Corporate value spells out and why selectPersonId reads "Empty if sign
    by Body Corporate". CR's schema gate takes any string up to 500 chars, so
    this is a wrong value CR ACCEPTS."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(
            graph(), year=2026,
            signatory={"name": "Get Started HK Limited",
                       "capacity": "Company Secretary", "person_id": None,
                       "date": None, "is_corporate": True},
        )
    assert any("Company Secretary" in p and "Body Corporate" in p
               for p in exc.value.problems)


def test_an_individual_may_not_sign_with_a_body_corporate_capacity():
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(
            graph(), year=2026,
            signatory={"name": "CHAN TAI MAN",
                       "capacity": "Director of the Company Secretary "
                                   "(Body Corporate)",
                       "person_id": "A1234567", "date": None},
        )
    assert any("Director" in p for p in exc.value.problems)


def test_a_capacity_that_is_in_neither_sheet_is_a_mapping_error():
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(
            graph(), year=2026,
            signatory={"name": "CHAN TAI MAN", "capacity": "Chief Executive",
                       "person_id": "A1234567", "date": None},
        )
    problem = next(p for p in exc.value.problems if "Chief Executive" in p)
    # Name the valid values -- CR's API is open six hours a day and a fault the
    # user cannot act on costs a round trip.
    assert "Authorized Representative" in problem


def test_a_derived_body_corporate_secretary_refuses_to_invent_a_capacity():
    """The mapper used to emit "Company Secretary" — an Individual value — for
    GSHK, a body corporate. Which of CR's 15 Body Corporate values GSHK signs
    under depends on who at GSHK signs; that is a business question, and the
    mapper's job is to make the invalid value impossible, not to pick."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(graph(), year=2026)
    problem = next(p for p in exc.value.problems if "Get Started HK" in p)
    assert "signatory=" in problem and "Body Corporate" in problem


def test_a_body_corporate_signing_with_a_body_corporate_capacity_is_accepted():
    data = nar1_mapper.map_entity(
        graph(), year=2026,
        signatory={"name": "Get Started HK Limited",
                   "capacity": "Company Secretary of the Company Secretary "
                               "(Body Corporate)",
                   "person_id": None, "date": None, "is_corporate": True},
    )
    assert data["selectCapacityDesc"] == ("Company Secretary of the Company "
                                          "Secretary (Body Corporate)")
    assert "selectPersonId" not in data


def test_a_return_with_nobody_to_sign_it_is_a_mapping_error():
    """Never emit the return with the declaration silently absent -- that is an
    unsigned statutory filing CR's schema gate happily accepts."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        nar1_mapper.map_entity(graph(secretaries=[]), year=2026)
    assert any("signatory" in p for p in exc.value.problems)


def test_a_private_company_declares_schedule_1_only():
    data = mapped(graph())
    assert data["shareholderListedInSch1"] == "Y"
    assert data["shareholderListedInSch2"] == "N"
    assert data["shareholderListedInCdrom"] == "N"


# ---- addresses -------------------------------------------------------------

def test_maps_a_registered_address_onto_crs_five_address_lines():
    data = mapped(graph())
    assert data["roAddr"] == {
        "addrLangInd": "E",
        "flatFlrBlk": "Flat A",
        "bldg": "Test Tower",
        "stEstLotVlg": "1 Test Street",
        "dstCtyStatePostal": "CENTRAL",
        "ctryRegion": "HKG",
    }


def test_country_becomes_a_three_letter_region_code():
    """ctryRegion is max 4 characters. "Hong Kong" would be truncated silently
    by CR or rejected; HKG is what CR's own example sends."""
    g = graph(registered_address={**ADDR, "country": "Hong Kong"})
    assert mapped(g)["roAddr"]["ctryRegion"] == "HKG"


def test_a_blank_country_is_hkg_on_the_must_be_hkg_nodes():
    """Three nodes carry the schema remark "Region. Must be HKG": roAddr,
    indSec/stdAddress and corpSec/stdAddress. A blank there is Hong Kong, and
    nowhere else is.

    The graph here is the shape nar1_source actually produces: a
    `company_secretaries` row with NO `corporate_address` (the loader never sets
    that key for secretaries — there is no corporate_entity_id on that table),
    so the GSHK secretary reuses the very same registered-office dict."""
    g = graph(registered_address={**ADDR, "country": None})
    data = mapped(g)
    assert data["roAddr"]["ctryRegion"] == "HKG"
    assert data["corpSecList"][0]["stdAddress"]["ctryRegion"] == "HKG"


def test_the_registered_office_is_hkg_even_when_the_row_says_hong_kong():
    """Emitted unconditionally, not derived — the schema allows one value."""
    assert mapped(graph())["roAddr"]["ctryRegion"] == "HKG"


def test_a_registered_office_recorded_outside_hong_kong_is_a_problem():
    """nar1_schema.json, roAddr/ctryRegion: "Region. Must be HKG". A HK company
    whose registered office is recorded in Vietnam is a data error worth saying
    out loud — silently overwriting it with HKG files an address nobody holds,
    and CR accepts that."""
    g = graph(registered_address={**ADDR, "country": "VN"})
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HKG" in p and "registered office" in p
               for p in exc.value.problems)


def test_an_individual_secretarys_address_must_be_hkg():
    """indSecList/indSec/stdAddress/ctryRegion: "Region. Must be HKG" — the same
    remark as roAddr, and different from every director node."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "company_secretary", "is_current": True}],
        secretaries=[],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": {**ADDR, "country": "GB"}},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HKG" in p and "CHAN TAI MAN" in p for p in exc.value.problems)


def test_a_corporate_secretarys_address_must_be_hkg():
    """corpSecList/corpSec/stdAddress/ctryRegion: "Region. Must be HKG"."""
    g = graph(
        officers=[{"corporate_name": "OFFSHORE SEC LIMITED",
                   "party_type": "corporate", "role": "company_secretary",
                   "is_current": True,
                   "corporate_address": {**CORP_ADDR, "country": "SG"}}],
        secretaries=[],
        addresses={"a1": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HKG" in p and "OFFSHORE SEC LIMITED" in p
               for p in exc.value.problems)


def test_a_directors_address_is_not_forced_to_hkg():
    """The guard that the HKG rule does not leak onto the nodes whose remark is
    "Refer to Country sheet" — most NAR1 directors do live abroad."""
    assert director_region("VN") == "VNM"


def test_a_blank_country_on_a_residential_address_is_a_problem_not_hong_kong():
    """A UK-resident director with a null country was silently filed as
    resident in Hong Kong -- wrong, and accepted by CR."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": {**ADDR, "country": None}},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("country" in p and "CHAN TAI MAN" in p for p in exc.value.problems)


def test_an_unmapped_country_is_a_mapping_error_not_a_guess():
    """Inventing a code produces a document CR rejects AFTER the fee is taken."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(resident_in("Freedonia"))
    assert any("Freedonia" in p for p in exc.value.problems)


# ---- CR's Country & Region vocabulary ---------------------------------------
#
# The old table keyed on 38 English names. DEV stores ISO alpha-2 and does so in
# 100% of its 8,027 address rows, so 87% of real addresses could not be filed at
# all. Every test below uses a shape DEV actually holds, not "Hong Kong".

def resident_in(country):
    """A graph whose one director lives in `country`.

    A director's stdAddress is the free-country node — CR's remark there is
    "Refer to Country sheet for Country code". roAddr and both secretary
    addresses are the "Must be HKG" nodes and cannot carry another country.
    """
    return graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": {**ADDR, "country": country}},
    )


def director_region(country):
    return mapped(resident_in(country))["indDirList"][0]["stdAddress"]["ctryRegion"]


#: (CR code, ISO alpha-2) for every row whose two columns do NOT start with the
#: same letter. Enumerated from the committed table and checked to be exhaustive
#: by the test below, so a row leaving or joining this set is a failure either
#: way. Ten rows, and each is a real historical naming divergence:
_FIRST_LETTER_EXCEPTIONS = {
    ("ATF", "TF"),   # FRENCH SOUTHERN TERRITORIES / Terres australes françaises
    ("COM", "KM"),   # COMOROS / Komori
    ("CYM", "KY"),   # CAYMAN ISLANDS
    ("GBR2", "JE"),  # JERSEY      — CR's own code. ISO would say JEY
    ("GBR3", "IM"),  # ISLE OF MAN — CR's own code. ISO would say IMN
    ("MYT", "YT"),   # MAYOTTE
    ("PRK", "KP"),   # DEMOCRATIC PEOPLE'S REPUBLIC OF KOREA / Korea, North
    ("SGS", "GS"),   # SOUTH GEORGIA AND THE SOUTH SANDWICH ISLANDS
    ("SPM", "PM"),   # SAINT PIERRE AND MIQUELON
    ("SRB", "RS"),   # SERBIA / Republika Srbija
}
# GBR1/GG (GUERNSEY) is deliberately absent — both columns start with G, so it
# obeys the rule despite being the third of CR's invented codes.


def test_the_country_table_still_carries_all_250_rows():
    """Unconditional row count — no workbook needed, so it holds in every
    environment.

    _build() refuses duplicate codes, duplicate alpha-2s and colliding
    descriptions, but it counts nothing. A table truncated by a bad merge or a
    half-pasted block is internally consistent and passes every other test in
    this file, then silently starts refusing countries at filing time.
    """
    assert len(cr_vocabularies._COUNTRY_ROWS) == 250
    assert len(cr_vocabularies.CR_COUNTRY_CODES) == 250
    assert len(cr_vocabularies.ALPHA2_TO_CR_CODE) == 250


def test_every_alpha_2_shares_a_first_letter_with_its_cr_code_bar_ten():
    """A structural rule over the whole table, not a sample of it.

    The 13 enumerated alpha-2 cases above catch an IN/IT-style swap, but only
    for the rows they name. This catches a swap of ANY two rows' alpha-2 columns
    whose first letters differ — the internally-consistent corruption _build()
    cannot see, which would file Sweden as Spain. Unconditional: it needs
    neither the workbook nor openpyxl.
    """
    off = {
        (code, alpha2)
        for code, _english, alpha2 in cr_vocabularies._COUNTRY_ROWS
        if alpha2[0] != code[0]
    }
    assert off == _FIRST_LETTER_EXCEPTIONS


def test_the_committed_country_table_is_crs_worksheet_row_for_row():
    """The data file is a transcription of CR's sheet, so prove it IS one.

    openpyxl is already a dev dependency and the workbook is committed under
    tests/fixtures, so this runs unconditionally — including in CI, which is the
    only place a transcription typo would otherwise reach a chargeable filing.
    """
    import openpyxl

    ws = openpyxl.load_workbook(WORKSHEET, read_only=True,
                                data_only=True)["Country & Region"]
    rows = [r for r in ws.iter_rows(values_only=True) if r and r[0]]
    theirs = {str(r[0]).strip(): str(r[1]).strip() for r in rows[1:]}

    assert len(theirs) == 250
    assert cr_vocabularies.CR_COUNTRY_CODES == theirs


def test_every_alpha_2_target_is_a_code_cr_actually_carries():
    """A typo in the alpha-2 table would ship a code CR rejects — after the fee.

    Nothing may resolve to a code outside CR's 250, and no two countries may
    claim the same one.
    """
    targets = cr_vocabularies.ALPHA2_TO_CR_CODE
    assert set(targets) and set(targets.values()) <= set(
        cr_vocabularies.CR_COUNTRY_CODES)
    assert len(set(targets.values())) == len(targets)
    for alias_target in cr_vocabularies._ALIASES.values():
        assert alias_target in cr_vocabularies.CR_COUNTRY_CODES


def test_the_crown_dependencies_use_crs_codes_not_isos():
    """CR invented GBR1/GBR2/GBR3 — which is why ctryRegion is max_length 4.
    pycountry and friends emit GGY/JEY/IMN and CR rejects all three."""
    assert director_region("GG") == "GBR1"
    assert director_region("JE") == "GBR2"
    assert director_region("IM") == "GBR3"
    assert cr_vocabularies.NON_ISO_COUNTRY_CODES == {"GBR1", "GBR2", "GBR3"}


@pytest.mark.parametrize("stored,code", [
    ("VN", "VNM"),   # 661 DEV rows
    ("AE", "ARE"),   # 532
    ("NL", "NLD"),   # 448
    ("GB", "GBR"),   # 412
    ("TH", "THA"),   # 379
    ("PH", "PHL"),   # 334
    ("IN", "IND"),   # 242
    ("CN", "CHN"),   # 235
    ("IT", "ITA"),   # 231
    ("AU", "AUS"),   # 221
    ("DE", "DEU"),   # 212
    ("HK", "HKG"),   # 819
    ("CH", "CHE"),   # Switzerland, NOT China — the classic alpha-2 trap
])
def test_an_iso_alpha_2_country_resolves(stored, code):
    """Every alpha-2 DEV stores in volume, pinned to the code CR expects.

    _build() already refuses a duplicated alpha-2, so a single mistyped cell
    cannot load; and every real ISO alpha-2 is claimed by some row, so a typo
    can only land on a string no address record contains. The one hole those
    two facts leave is a SWAP of two rows' alpha-2 columns, which is internally
    consistent and would file Sweden as Spain. These cases close that hole for
    the rows that actually carry production volume.
    """
    assert director_region(stored) == code


@pytest.mark.parametrize("stored", ["HKG", "VNM", "GBR1", "hkg"])
def test_a_value_that_is_already_a_cr_code_passes_through(stored):
    assert director_region(stored) == stored.upper()


@pytest.mark.parametrize("stored", ["VIET NAM", "Vietnam", "viet nam",
                                    " Viet  Nam "])
def test_crs_own_english_description_resolves_however_it_is_spaced(stored):
    """CR writes "VIET NAM"; G-FlowDesk writes "Vietnam". Both sides normalise."""
    assert director_region(stored) == "VNM"


@pytest.mark.parametrize("stored,code", [
    ("Hong Kong", "HKG"), ("hongkong", "HKG"), ("uk", "GBR"),
    ("britain", "GBR"), ("usa", "USA"), ("us", "USA"), ("bvi", "VGB"),
    ("prc", "CHN"), ("macau", "MAC"), ("macao", "MAC"),
    ("south korea", "KOR"), ("korea", "KOR"),
    ("united states of america", "USA"), ("new zealand", "NZL"),
])
def test_the_hand_written_aliases_still_resolve(stored, code):
    """Older rows may still carry these, so replacing the table must not start
    refusing values that used to map."""
    assert director_region(stored) == code


@pytest.mark.parametrize("stored", ["Freedonia", "GB-ENG", "TW-CH", "XX"])
def test_a_country_cr_has_no_code_for_is_still_a_problem(stored):
    """Fail-loud is the whole point: GB-ENG and TW-CH are real DEV values and
    neither is a country code. Guessing GBR/TWN would file a value CR accepts
    off data nobody validated."""
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(resident_in(stored))
    assert any(stored in p for p in exc.value.problems)


# ---- officers --------------------------------------------------------------

def test_an_individual_director_lands_in_inddirlist():
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    data = mapped(g)
    assert len(data["indDirList"]) == 1
    d = data["indDirList"][0]
    assert d["indvEngSname"] == "CHAN"
    assert d["indvEngOname"] == "TAI MAN"
    assert d["indvChiName"] == "陳大文"
    assert d["dirInd"] == "Y"


def test_a_corporate_director_lands_in_corpdirlist():
    g = graph(
        officers=[{"corporate_name": "HOLDCO LIMITED", "party_type": "corporate",
                   "role": "director", "is_current": True,
                   "corporate_address": CORP_ADDR}],
        addresses={"a1": ADDR},
    )
    data = mapped(g)
    assert data["corpDirList"][0]["corpEngName"] == "HOLDCO LIMITED"
    assert "indDirList" not in data


def test_a_corporate_officer_files_its_own_address_not_the_filers():
    """Migration 007 gave entity_officers a corporate_entity_id FK, so the real
    address is reachable. Substituting the filing company's registered office
    files a WRONG address on a statutory return -- and CR accepts it."""
    g = graph(
        officers=[{"corporate_name": "HOLDCO LIMITED", "party_type": "corporate",
                   "role": "director", "is_current": True,
                   "corporate_address": CORP_ADDR}],
        addresses={"a1": ADDR},
    )
    addr = mapped(g)["corpDirList"][0]["stdAddress"]
    assert addr["bldg"] == "Corp Tower"
    assert addr["dstCtyStatePostal"] == "WAN CHAI"


def test_a_corporate_officers_br_number_and_chinese_name_are_filed():
    """_corporate's br_no / name_zh parameters were dead -- no caller passed
    them -- while the worksheet defines corpBrNo and corpChiName and CR's own
    example sends both."""
    g = graph(
        officers=[{"corporate_name": "HOLDCO LIMITED", "party_type": "corporate",
                   "role": "director", "is_current": True,
                   "corporate_address": CORP_ADDR,
                   "corporate_br_no": "00000002",
                   "corporate_name_zh": "控股有限公司"}],
        addresses={"a1": ADDR},
    )
    corp = mapped(g)["corpDirList"][0]
    assert corp["corpBrNo"] == "00000002"
    assert corp["corpChiName"] == "控股有限公司"


def test_a_corporate_officer_with_no_address_is_a_mapping_error():
    g = graph(
        officers=[{"corporate_name": "HOLDCO LIMITED", "party_type": "corporate",
                   "role": "director", "is_current": True}],
        addresses={"a1": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HOLDCO LIMITED" in p for p in exc.value.problems)


def test_a_reserve_director_lands_in_resdirlist():
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "reserve_director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    data = mapped(g)
    assert len(data["resDirList"]) == 1
    assert "indDirList" not in data


def test_a_resigned_officer_is_excluded():
    """The annual return states the officers AT the return date. A resigned
    director filed as current is a false statutory declaration."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": False}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    assert "indDirList" not in mapped(g)


def test_the_gshk_secretary_is_a_corporate_secretary_with_its_tcsp_number():
    g = graph(
        secretaries=[{"is_gshk": True, "secretary_name": "Get Started HK Limited",
                      "tcsp_number": "TC000807", "is_current": True}],
        addresses={"a1": ADDR},
    )
    sec = mapped(g)["corpSecList"][0]
    assert sec["corpEngName"] == "Get Started HK Limited"
    assert sec["corpTcspNo"] == "TC000807"


def test_a_gshk_corporate_secretary_falls_back_to_the_registered_office():
    """`company_secretaries` has no corporate_entity_id (migration 007 put that
    FK on entity_officers / shareholdings / beneficial_owners only), so a
    corporate secretary has no address of its own to file. For the GSHK
    secretary the filing company's registered office IS GSHK's own address, by
    construction — GSHK provides it. That is the ONE place the filer's address
    may stand in for a corporate party's, and it must not widen."""
    sec = mapped(graph())["corpSecList"][0]
    assert sec["corpEngName"] == "Get Started HK Limited"
    assert sec["stdAddress"]["bldg"] == "Test Tower"       # the filer's RO


def test_a_non_gshk_corporate_secretary_with_no_address_is_a_mapping_error():
    """A client that keeps its own registered office would be misfiled if the
    fallback applied to any body corporate — Critical 3's rule holds everywhere
    except the GSHK secretary."""
    g = graph(secretaries=[{"secretary_name": "OTHER SEC LIMITED",
                            "is_gshk": False, "is_current": True}],
              addresses={"a1": ADDR})
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("OTHER SEC LIMITED" in p and "no address" in p
               for p in exc.value.problems)


def test_a_secretary_recorded_in_both_tables_is_emitted_once():
    """officer_role includes 'company_secretary', so the same individual can be
    an entity_officers row AND a company_secretaries row. Two indSec entries
    for one secretary is a false statutory return."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "company_secretary", "is_current": True}],
        secretaries=[{"person_id": "p1", "is_gshk": False, "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "hkid", "id_number": "A123456(7)",
                                    "is_primary": True}]},
    )
    assert len(mapped(g)["indSecList"]) == 1


def test_a_corporate_secretary_in_both_tables_is_emitted_once_with_its_tcsp():
    """Dedup falls back to a normalised name for rows with no person_id, and
    the company_secretaries row -- the one carrying the TCSP number -- wins."""
    g = graph(
        officers=[{"corporate_name": "Get Started HK  Limited",
                   "party_type": "corporate", "role": "company_secretary",
                   "is_current": True, "corporate_address": CORP_ADDR}],
        secretaries=[{"is_gshk": True, "secretary_name": "GET STARTED HK LIMITED",
                      "tcsp_number": "TC000807", "is_current": True}],
        addresses={"a1": ADDR},
    )
    secs = mapped(g)["corpSecList"]
    assert len(secs) == 1
    assert secs[0]["corpTcspNo"] == "TC000807"


def test_hkid_is_sent_as_the_partial_number_cr_asks_for():
    """indvHkidNo on a NAR1 is the PARTIAL HKID: leading letters plus the first
    three digits, at most 5 characters. Verified live 2026-08-21 -- CR rejects
    a full 8-character number with "HKID No. length must be at most 5", and
    CR's own workbook has a separate "Partial HKID" column feeding this field.
    Sending the full number would also disclose more than CR asks for."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "hkid", "id_number": "A123456(7)",
                                    "is_primary": True}]},
    )
    assert mapped(g)["indDirList"][0]["indvHkidNo"] == "A123"


def test_a_china_id_is_a_problem_not_a_director_filed_without_any_id():
    """id_document_type is ENUM('hkid','passport','china_id','other') and the
    NAR1 carries only indvHkidNo / indvPptNo — no node in nar1_schema.json
    mentions a PRC identity card. A PRC director was being filed with NO
    identity number at all and nothing appended to problems."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "china_id",
                                    "id_number": "440101199001011234",
                                    "is_primary": True}]},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("china_id" in p and "CHAN TAI MAN" in p for p in exc.value.problems)


def test_a_china_id_alongside_a_passport_files_the_passport_quietly():
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [
            {"id_type": "china_id", "id_number": "440101199001011234",
             "is_primary": True},
            {"id_type": "passport", "id_number": "E12345678",
             "issuing_country": "China"},
        ]},
    )
    assert mapped(g)["indDirList"][0]["indvPptNo"] == "E12345678"


def test_a_passport_row_with_no_number_does_not_crash_the_mapper():
    """`passport["id_number"]` was the one unguarded subscript left in a helper
    that reads every other column with .get(). A null column is a KeyError —
    an unhandled 500 on the prepare endpoint rather than a fault the user can
    read. It is now a reported problem instead: CR requires a partial identity
    number for every individual, so a passport row with no number cannot be
    filed and must be said out loud."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "passport",
                                    "issuing_country": "Singapore"}]},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("passport" in p for p in exc.value.problems)


def test_an_hkid_that_yields_no_partial_number_is_caught_in_the_mapper():
    """Two letters plus three digits IS valid ("XA123" is one of CR's own two
    examples), so the guard is on the SHAPE, not merely on length."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "hkid", "id_number": "!!!",
                                    "is_primary": True}]},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HKID" in p for p in exc.value.problems)


def test_a_passport_holder_gets_a_number_and_an_issuing_country():
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
        identity_documents={"p1": [{"id_type": "passport", "id_number": "X1234567",
                                    "issuing_country": "Singapore",
                                    "is_primary": True}]},
    )
    d = mapped(g)["indDirList"][0]
    assert d["indvPptNo"] == "X1234567"
    assert d["indvPptIssCtry"] == "SGP"
    assert "indvHkidNo" not in d


# ---- share capital and Schedule 1 ------------------------------------------

def test_each_share_class_becomes_one_sharecapital_entry():
    # Each class needs a holder for its full issued count: every loaded class is
    # reconciled against Schedule 1, so issued shares with no member on record
    # are now (correctly) a MappingError rather than a silent filing.
    g = graph(share_classes=[
        {"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
         "total_issued": 1000, "total_paid": 1000},
        {"id": "sc2", "class_name": "Preference", "currency": "CAD",
         "total_issued": 2000, "total_paid": 2000},
    ], shareholdings=[
        {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
         "shares_held": 1000, "is_current": True},
        {"share_class_id": "sc2", "person_id": "p1", "party_type": "individual",
         "shares_held": 2000, "is_current": True},
    ], persons={"p1": person()}, addresses={"a1": ADDR, "a2": ADDR})
    caps = mapped(g)["shareCapitals"]
    assert len(caps) == 2
    assert caps[0] == {
        "clsOfShares": "Ordinary", "currency": "HKD",
        "noOfShareIssuedOnThisCls": 1000, "issuedCapital": 1000,
        "paidUpCapital": 1000,
    }


def test_schedule_1_groups_shareholders_under_their_share_class():
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1000}],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 100, "is_current": True},
            {"share_class_id": "sc1", "corporate_name": "TEST COMPANY LIMITED",
             "party_type": "corporate", "shares_held": 900, "is_current": True,
             "corporate_address": CORP_ADDR},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    share = mapped(g)["schedule1"]["shares"][0]
    assert share["clsOfShares"] == "Ordinary"
    grps = share["shareHolderGrps"]
    assert [g_["sharesAlloted"] for g_ in grps] == [100, 900]
    assert grps[0]["allotteeRec"][0]["allotteeType"] == "I"
    assert grps[1]["allotteeRec"][0]["allotteeType"] == "C"


def test_shtype_is_1_for_a_sole_individual_shareholder():
    """shType is SHAREHOLDER TYPE -- "1" sole, "2" joint. Nothing to do with
    whether the shares are paid up."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 100, "total_paid": 100}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": 100,
                        "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    grp = mapped(g)["schedule1"]["shares"][0]["shareHolderGrps"][0]
    assert grp["shType"] == "1"


def test_shtype_is_1_for_a_sole_corporate_shareholder():
    """CR's example lines 172-191: a shType=1 group whose only allottee is
    `C`. So "1" means ONE holder, not "natural person" -- filing a corporate
    sole shareholder as "2" (or as "Individual Shareholder") is a
    self-contradicting return CR's schema happily accepts."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 900, "total_paid": 900}],
        shareholdings=[{"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
                        "party_type": "corporate", "shares_held": 900,
                        "is_current": True, "corporate_address": ADDR}],
        addresses={"a1": ADDR},
    )
    grp = mapped(g)["schedule1"]["shares"][0]["shareHolderGrps"][0]
    assert grp["shType"] == "1"
    assert grp["allotteeRec"][0]["allotteeType"] == "C"


def test_a_corporate_shareholder_files_its_own_address_not_the_filers():
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 900, "total_paid": 900}],
        shareholdings=[{"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
                        "party_type": "corporate", "shares_held": 900,
                        "is_current": True, "corporate_address": CORP_ADDR}],
        addresses={"a1": ADDR},
    )
    grp = mapped(g)["schedule1"]["shares"][0]["shareHolderGrps"][0]
    assert grp["allotteeRec"][0]["allotteeAddr"]["bldg"] == "Corp Tower"


def test_a_corporate_shareholder_with_no_address_is_a_mapping_error():
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 900, "total_paid": 900}],
        shareholdings=[{"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
                        "party_type": "corporate", "shares_held": 900,
                        "is_current": True}],
        addresses={"a1": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("HOLDCO LIMITED" in p for p in exc.value.problems)


def test_shtype_ignores_amount_paid_entirely():
    """amount_paid is numeric(20,4) MONEY and shares_held is a share COUNT, so
    comparing them is a unit mismatch that coincides only for HKD-1-par shares
    -- and payment state is not what shType means in the first place."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": 1000,
                        "amount_paid": 1, "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    grp = mapped(g)["schedule1"]["shares"][0]["shareHolderGrps"][0]
    assert grp["shType"] == "1"


def test_a_holding_whose_share_class_was_not_loaded_is_a_mapping_error():
    """The emit loop iterates share_classes and picks matching holdings, so a
    holding pointing at a class that is not loaded was dropped in silence -- a
    shareholder disappearing from the statutory register of members."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 100, "total_paid": 100}],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 100, "is_current": True},
            {"share_class_id": "sc9", "person_id": "p1", "party_type": "individual",
             "shares_held": 500, "is_current": True},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("sc9" in p for p in exc.value.problems)


def test_shares_allotted_short_of_shares_issued_is_a_mapping_error():
    """A class with 1000 issued and 900 allotted filed silently."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1000}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": 900,
                        "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("900" in p and "1000" in p for p in exc.value.problems)


def test_a_joint_block_counts_once_against_shares_issued():
    """Two rows holding one block jointly is 2000 allotted, not 4000."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 2000, "total_paid": 2000}],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 2000, "is_current": True, "joint_group_id": "j1"},
            {"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
             "party_type": "corporate", "shares_held": 2000, "is_current": True,
             "joint_group_id": "j1", "corporate_address": CORP_ADDR},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    assert mapped(g)["schedule1"]["shares"][0][
        "shareHolderGrps"][0]["sharesAlloted"] == 2000


def test_a_fractional_share_count_is_a_mapping_error_not_a_truncation():
    """total_issued is numeric(20,4) and every CR field here is an Integer, so
    int() was silently truncating a number CR cannot represent."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": "1000.5000", "total_paid": 1000}],
        addresses={"a1": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("1000.5000" in p for p in exc.value.problems)


def test_an_unparseable_holding_size_is_reported_once_not_twice():
    """The class total and the group's sharesAlloted both parsed the SAME
    shares_held value, so one unparseable number produced the identical problem
    twice. A duplicated fault list is a user fixing the same field twice against
    an API that is open six hours a day."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 100, "total_paid": 100}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": "not-a-number",
                        "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert len([p for p in exc.value.problems if "is not a number" in p]) == 1


def test_a_share_class_with_no_current_holder_at_all_is_still_reconciled():
    """The issued-vs-allotted reconciliation used to run only for classes that
    emitted a group, so a class with 1000 issued and nobody holding it produced
    an empty Schedule 1 while shareCapitals still declared the 1000 — a
    thousand shares belonging to nobody, filed silently."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1000}],
        shareholdings=[],
        addresses={"a1": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("accounts for 0" in p and "1000 issued" in p
               for p in exc.value.problems)


def test_a_former_shareholding_is_excluded():
    """The former holding stays out of the emitted groups — and because it does,
    the class accounts for 0 of its 1000 issued shares, which is the same defect
    the reconciliation exists to catch. "accounts for 0", not "accounts for 100",
    is what proves the exclusion."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1000}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": 100,
                        "is_current": False}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("accounts for 0" in p and "1000 issued" in p
               for p in exc.value.problems)


# ---- the round trip --------------------------------------------------------

def _cardinality(xml: str) -> dict[str, int]:
    """Count every repeating wrapper's children. Silent data loss hides HERE:
    a mapper that drops the second director still produces valid XML."""
    return {
        tag: len(re.findall(rf"<cr:{tag}>", xml))
        for tag in ("shareCapital", "indSec", "corpSec", "indDir", "corpDir",
                    "resDir", "share", "shareHolderGrp", "allottee")
    }


def test_the_mapped_dict_builds_without_a_validation_error():
    """nar1.validate() is the schema gate. A mapper key CR does not know is a
    document CR rejects -- catch it here, not at a chargeable call."""
    g = graph(
        officers=[{"person_id": "p1", "party_type": "individual",
                   "role": "director", "is_current": True}],
        secretaries=[{"is_gshk": True, "secretary_name": "Get Started HK Limited",
                      "tcsp_number": "TC000807", "is_current": True}],
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 1000, "total_paid": 1000}],
        shareholdings=[{"share_class_id": "sc1", "person_id": "p1",
                        "party_type": "individual", "shares_held": 1000,
                        "is_current": True}],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    data = mapped(g)
    assert nar1.validate(data) == []
    xml = nar1.build_nar1_xml(data)
    assert "<cr:brNo>00000001</cr:brNo>" in xml


def test_round_trip_cardinality_matches_crs_own_example():
    """Build the same company CR's example describes and assert the repeating
    wrappers come out with the same counts. This is the assertion spec §5 BE-1
    calls for: cardinality on repeating wrappers is where data loss hides."""
    g = graph(
        officers=[
            {"person_id": "p1", "party_type": "individual", "role": "director",
             "is_current": True},
            {"corporate_name": "Test Company Limited", "party_type": "corporate",
             "role": "director", "is_current": True,
             "corporate_address": CORP_ADDR},
            {"person_id": "p1", "party_type": "individual",
             "role": "reserve_director", "is_current": True},
            {"person_id": "p1", "party_type": "individual",
             "role": "company_secretary", "is_current": True},
        ],
        secretaries=[{"is_gshk": True, "secretary_name": "TEST COMPANY LIMITED",
                      "tcsp_number": "TC000807", "is_current": True}],
        share_classes=[
            {"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
             "total_issued": 1000, "total_paid": 1000},
            {"id": "sc2", "class_name": "Preference", "currency": "CAD",
             "total_issued": 2000, "total_paid": 2000},
        ],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 100, "is_current": True},
            {"share_class_id": "sc1", "corporate_name": "TEST COMPANY LIMITED",
             "party_type": "corporate", "shares_held": 900, "is_current": True,
             "corporate_address": CORP_ADDR},
            # CR's Preference class is ONE group of 2000 shares held JOINTLY by
            # an individual and a body corporate -- shType 2, two allottees.
            {"share_class_id": "sc2", "person_id": "p1", "party_type": "individual",
             "shares_held": 2000, "is_current": True, "joint_group_id": "j1"},
            {"share_class_id": "sc2", "corporate_name": "Test Company Limited",
             "party_type": "corporate", "shares_held": 2000, "is_current": True,
             "joint_group_id": "j1", "corporate_address": CORP_ADDR},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    ours = _cardinality(nar1.build_nar1_xml(mapped(g, year=2020)))
    theirs = _cardinality(SAMPLE.read_text(encoding="utf8"))

    # `allottee` is in this loop. It was computed and NOT asserted before, and
    # the gap it hid (ours 3, CR's 4) was the joint-shareholder concept.
    for tag in ("shareCapital", "indSec", "corpSec", "indDir", "corpDir",
                "resDir", "share", "shareHolderGrp", "allottee"):
        assert ours[tag] == theirs[tag], (
            f"{tag}: built {ours[tag]}, CR's example has {theirs[tag]}"
        )


def test_joint_holders_are_one_group_of_shtype_2_with_one_allottee_each():
    """CR's Preference class (lines 197-231): one shareHolderGrp, shType 2, two
    allottees. Splitting joint holders into separate sole groups misdeclares
    the register -- and CR's schema accepts it."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 2000, "total_paid": 2000}],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 2000, "is_current": True, "joint_group_id": "j1"},
            {"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
             "party_type": "corporate", "shares_held": 2000, "is_current": True,
             "joint_group_id": "j1", "corporate_address": CORP_ADDR},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    grps = mapped(g)["schedule1"]["shares"][0]["shareHolderGrps"]
    assert len(grps) == 1
    assert grps[0]["shType"] == "2"
    assert grps[0]["sharesAlloted"] == 2000
    assert [a["allotteeType"] for a in grps[0]["allotteeRec"]] == ["I", "C"]


def test_joint_holders_disagreeing_on_the_block_size_is_a_mapping_error():
    """Joint holders hold ONE block jointly. Two different numbers means the
    register is wrong, and picking either would file a number nobody recorded."""
    g = graph(
        share_classes=[{"id": "sc1", "class_name": "Ordinary", "currency": "HKD",
                        "total_issued": 2000, "total_paid": 2000}],
        shareholdings=[
            {"share_class_id": "sc1", "person_id": "p1", "party_type": "individual",
             "shares_held": 2000, "is_current": True, "joint_group_id": "j1"},
            {"share_class_id": "sc1", "corporate_name": "HOLDCO LIMITED",
             "party_type": "corporate", "shares_held": 1500, "is_current": True,
             "joint_group_id": "j1", "corporate_address": CORP_ADDR},
        ],
        persons={"p1": person()},
        addresses={"a1": ADDR, "a2": ADDR},
    )
    with pytest.raises(nar1_mapper.MappingError) as exc:
        mapped(g)
    assert any("joint" in p for p in exc.value.problems)


def test_the_mapper_does_no_io():
    """The riskiest form logic in the system must be testable without CR and
    without Supabase -- the TEST form APIs are open 30 hours a week."""
    import inspect
    src = inspect.getsource(nar1_mapper)
    assert "supabase" not in src.lower()
